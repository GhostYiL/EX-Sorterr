import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox, Checkbutton, IntVar, ttk, scrolledtext, StringVar

try:
    from tabulate import tabulate
    USE_TABULATE = True
except ImportError:
    USE_TABULATE = False

def extract_first_number(text):
    if not isinstance(text, str):
        return float('inf')
    match = re.search(r'\d+', str(text))
    return int(match.group()) if match else float('inf')

LANGS = {
    "en": {
        "title": "Excel Sorter",
        "header": "File has header row",
        "select": "Select Excel File",
        "save": "Save Sorted File",
        "preview": "[PREVIEW] (first 10 rows):",
        "sorted": "[SORTED PREVIEW] (first 10 rows):",
        "by": "by - Mist",
        "select_file": "Select Excel file",
        "save_file": "Save sorted file as",
        "success": "Success",
        "saved": "Sorted file saved as:\n{}",
        "error": "ERROR",
        "could_not_load": "Could not load file:\n{}",
        "could_not_save": "Could not save file:\n{}",
        "no_data": "No Data",
        "please_load": "Please load a file first.",
        "lang": "Language",
        "sort_col": "Sort by column:",
        "cancel": "Cancel",
        "sheet": "Sheet:",
        "all_sheets": "All Sheets",
        "wrap_text_column": "Wrap text in column:",
    },
    "ru": {
        "title": "Сортировка Excel",
        "header": "В файле есть заголовок",
        "select": "Выбрать Excel файл",
        "save": "Сохранить",
        "preview": "[ПРОСМОТР] (первые 10 строк):",
        "sorted": "[ОТСОРТИРОВАНО] (первые 10 строк):",
        "by": "by - Mist",
        "select_file": "Выберите файл Excel",
        "save_file": "Сохранить отсортированный файл как",
        "success": "Успех",
        "saved": "Отсортированный файл сохранён:\n{}",
        "error": "ОШИБКА",
        "could_not_load": "Не удалось загрузить файл:\n{}",
        "could_not_save": "Не удалось сохранить файл:\n{}",
        "no_data": "Нет данных",
        "please_load": "Сначала выберите файл.",
        "lang": "Язык",
        "sort_col": "Сортировать по столбцу:",
        "cancel": "Отмена",
        "sheet": "Лист:",
        "all_sheets": "Все листы",
        "wrap_text_column": "Перенос текста в столбце:",
    }
}

def autosize_and_wrap_column(filename, wrap_col_name=None, wrap_width=40):
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
    wb = openpyxl.load_workbook(filename)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for ws in wb.worksheets:
        wrap_idx = None
        if wrap_col_name:
            for idx, cell in enumerate(ws[1], 1):
                if str(cell.value).strip() == str(wrap_col_name).strip():
                    wrap_idx = idx
                    break

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(value))
            if wrap_idx and col[0].column == wrap_idx:
                ws.column_dimensions[col_letter].width = wrap_width
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                ws.column_dimensions[col_letter].width = min(30, max_length + 2)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        # Titles
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
    wb.save(filename)

class ExcelSorterApp:
    def __init__(self, root):
        self.root = root
        self.lang = StringVar(value="ru")
        self.df = None
        self.df_sorted = None
        self.file_path = None
        self.sheet_names = []
        self.sheet_var = StringVar()
        self.header_var = IntVar(value=0)
        self.sort_col_var = StringVar(value="0")
        self.col_names = []
        self.wrap_col_var = StringVar(value="")

        # Language selection
        lang_frame = tk.Frame(root)
        lang_frame.pack(anchor="ne", padx=10, pady=5)
        tk.Label(lang_frame, text=LANGS["en"]["lang"] + "/" + LANGS["ru"]["lang"] + ":").pack(side="left")
        lang_menu = ttk.Combobox(lang_frame, textvariable=self.lang, values=["en", "ru"], width=5, state="readonly")
        lang_menu.pack(side="left")
        lang_menu.bind("<<ComboboxSelected>>", self.update_language)

        self.widgets = {}

        self.root.title(LANGS[self.lang.get()]["title"])
        self.root.geometry("800x600")

        self.widgets['header'] = Checkbutton(root, text=LANGS[self.lang.get()]["header"], variable=self.header_var, command=self.on_header_toggle)
        self.widgets['header'].pack(anchor="w", padx=10, pady=5)

        wrap_frame = tk.Frame(root)
        wrap_frame.pack(anchor="w", padx=10, pady=3)
        self.widgets['wrap_label'] = tk.Label(wrap_frame, text=LANGS[self.lang.get()]["wrap_text_column"])
        self.widgets['wrap_label'].pack(side="left")
        self.wrap_col_combo = ttk.Combobox(wrap_frame, textvariable=self.wrap_col_var, state="readonly", width=25)
        self.wrap_col_combo.pack(side="left", padx=5)

        sort_frame = tk.Frame(root)
        sort_frame.pack(anchor="w", padx=10, pady=5)
        self.widgets['sheet_label'] = tk.Label(sort_frame, text=LANGS[self.lang.get()]["sheet"])
        self.widgets['sheet_label'].pack(side="left")
        self.sheet_combo = ttk.Combobox(sort_frame, textvariable=self.sheet_var, state="readonly", width=20)
        self.sheet_combo.pack(side="left", padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)

        self.widgets['sort_label'] = tk.Label(sort_frame, text=LANGS[self.lang.get()]["sort_col"])
        self.widgets['sort_label'].pack(side="left")
        self.sort_col_combo = ttk.Combobox(sort_frame, textvariable=self.sort_col_var, state="readonly", width=25)
        self.sort_col_combo.pack(side="left", padx=5)
        self.sort_col_combo.bind("<<ComboboxSelected>>", self.on_sort_col_change)

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=5)
        self.widgets['select'] = tk.Button(btn_frame, text=LANGS[self.lang.get()]["select"], command=self.load_and_preview)
        self.widgets['select'].pack(side="left", padx=5)
        self.widgets['save'] = tk.Button(btn_frame, text=LANGS[self.lang.get()]["save"], command=self.save_sorted_file, state="disabled")
        self.widgets['save'].pack(side="left", padx=5)
        self.widgets['cancel'] = tk.Button(btn_frame, text=LANGS[self.lang.get()]["cancel"], command=self.cancel_selection)
        self.widgets['cancel'].pack(side="left", padx=5)

        preview_frame = tk.Frame(root)
        preview_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.file_label = tk.Label(preview_frame, text="", fg="blue", anchor="w", font=("Segoe UI", 9, "italic"))
        self.file_label.pack(fill="x", pady=(0, 2))

        self.widgets['preview_label'] = tk.Label(preview_frame, text=LANGS[self.lang.get()]["preview"])
        self.widgets['preview_label'].pack(anchor="w")
        self.text_preview = scrolledtext.ScrolledText(preview_frame, height=10, width=110, font=("Consolas", 10))
        self.text_preview.pack(fill="both", expand=True)

        self.widgets['sorted_label'] = tk.Label(preview_frame, text=LANGS[self.lang.get()]["sorted"])
        self.widgets['sorted_label'].pack(anchor="w", pady=(10,0))
        self.text_sorted = scrolledtext.ScrolledText(preview_frame, height=10, width=110, font=("Consolas", 10))
        self.text_sorted.pack(fill="both", expand=True)

        self.widgets['by'] = tk.Label(root, text=LANGS[self.lang.get()]["by"], fg="gray")
        self.widgets['by'].pack(side="bottom", pady=5)

        self.update_language()

    def update_language(self, event=None):
        lang = self.lang.get()
        self.root.title(LANGS[lang]["title"])
        self.widgets['header'].config(text=LANGS[lang]["header"])
        self.widgets['select'].config(text=LANGS[lang]["select"])
        self.widgets['save'].config(text=LANGS[lang]["save"])
        self.widgets['cancel'].config(text=LANGS[lang]["cancel"])
        self.widgets['preview_label'].config(text=LANGS[lang]["preview"])
        self.widgets['sorted_label'].config(text=LANGS[lang]["sorted"])
        self.widgets['by'].config(text=LANGS[lang]["by"])
        self.widgets['sort_label'].config(text=LANGS[lang]["sort_col"])
        self.widgets['sheet_label'].config(text=LANGS[lang]["sheet"])
        self.widgets['wrap_label'].config(text=LANGS[lang]["wrap_text_column"])

    def pretty_preview(self, df, highlight_col=None):
        if df is None:
            return ""
        df = df.head(10)
        if highlight_col is not None:
            df = df.copy()
            col_list = list(df.columns)
            if isinstance(highlight_col, int) and 0 <= highlight_col < len(col_list):
                col_list[highlight_col] = f"*{col_list[highlight_col]}*"
                df.columns = col_list
                df.iloc[:, highlight_col] = df.iloc[:, highlight_col].astype(str)
                for i in range(len(df)):
                    df.iloc[i, highlight_col] = f"> {df.iloc[i, highlight_col]} <"
        if USE_TABULATE:
            return tabulate(df, headers="keys", tablefmt="github", showindex=False)
        else:
            return df.to_string(index=False)

    def load_and_preview(self):
        lang = self.lang.get()
        self.file_path = filedialog.askopenfilename(title=LANGS[lang]["select_file"], filetypes=[("Excel files", "*.xlsx")])
        if not self.file_path:
            return
        import os
        self.file_label.config(text=os.path.basename(self.file_path))
        try:
            xl = pd.ExcelFile(self.file_path)
            self.sheet_names = xl.sheet_names
            if not self.sheet_names:
                messagebox.showerror(LANGS[lang]["error"], LANGS[lang]["no_data"])
                return
            all_sheets_label = LANGS[lang]["all_sheets"]
            # Format
            if lang == "ru":
                sheet_options = [all_sheets_label] + [f"Лист {i+1}" for i in range(len(self.sheet_names))] if len(self.sheet_names) > 1 else [f"Лист 1"]
            else:
                sheet_options = [all_sheets_label] + [f"Sheet {i+1}" for i in range(len(self.sheet_names))] if len(self.sheet_names) > 1 else [f"Sheet 1"]
            self.sheet_combo['values'] = sheet_options
            self.sheet_var.set(sheet_options[0])
        except Exception as e:
            messagebox.showerror(LANGS[lang]["error"], LANGS[lang]["could_not_load"].format(e))
            return
        self._load_and_show()
        self.widgets['save'].config(state="normal")

    def on_sheet_change(self, event=None):
        self._load_and_show()

    def _load_and_show(self):
        lang = self.lang.get()
        try:
            sheet = self.sheet_var.get()
            all_sheets_label = LANGS[lang]["all_sheets"]
            if not self.sheet_names:
                messagebox.showerror(LANGS[lang]["error"], LANGS[lang]["no_data"])
                return
            if sheet == all_sheets_label:
                sheet_name = self.sheet_names[0]
            else:
                idx = int(sheet.split()[-1]) - 1
                sheet_name = self.sheet_names[idx]
            if sheet_name not in self.sheet_names:
                messagebox.showerror(LANGS[lang]["error"], f"{LANGS[lang]['sheet']} '{sheet_name}' {LANGS[lang]['could_not_load'].split(':')[0].lower()}.")
                return
            self.df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0 if self.header_var.get() else None)
            if self.df.empty or len(self.df.columns) == 0:
                self.text_preview.config(state="normal")
                self.text_preview.delete(1.0, tk.END)
                self.text_preview.insert(tk.END, f"{LANGS[lang]['no_data']}: {sheet}")
                self.text_preview.config(state="disabled")
                self.text_sorted.config(state="normal")
                self.text_sorted.delete(1.0, tk.END)
                self.text_sorted.insert(tk.END, f"{LANGS[lang]['no_data']}: {sheet}")
                self.text_sorted.config(state="disabled")
                self.widgets['save'].config(state="disabled")
                return
            self.col_names = list(self.df.columns)
            # Sort columns
            col_options = [f"{i}: {name}" for i, name in enumerate(self.col_names)]
            self.sort_col_combo['values'] = col_options

            try:
                idx = int(self.sort_col_var.get().split(":")[0])
            except Exception:
                idx = 0
            if idx < 0 or idx >= len(self.col_names):
                idx = 0
            self.sort_col_var.set(col_options[idx])

            # Refresh the list
            self.wrap_col_combo['values'] = self.col_names
            if self.col_names:
                if self.wrap_col_var.get() not in self.col_names:
                    self.wrap_col_var.set(self.col_names[-1])

            self.text_preview.config(state="normal")
            self.text_preview.delete(1.0, tk.END)
            self.text_preview.insert(tk.END, self.pretty_preview(self.df, highlight_col=idx))
            self.text_preview.config(state="disabled")

            self.df_sorted = self.df.copy()
            sort_col = idx
            self.df_sorted['__sort_key__'] = self.df_sorted.iloc[:, sort_col].apply(extract_first_number)
            self.df_sorted = self.df_sorted.sort_values('__sort_key__').drop(columns='__sort_key__')

            self.text_sorted.config(state="normal")
            self.text_sorted.delete(1.0, tk.END)
            self.text_sorted.insert(tk.END, self.pretty_preview(self.df_sorted, highlight_col=idx))
            self.text_sorted.config(state="disabled")
            self.widgets['save'].config(state="normal")
        except Exception as e:
            self.widgets['save'].config(state="disabled")
            messagebox.showerror(LANGS[lang]["error"], LANGS[lang]["could_not_load"].format(e))

    def on_header_toggle(self):
        if self.file_path:
            self._load_and_show()

    def on_sort_col_change(self, event=None):
        if self.df is not None:
            self._load_and_show()

    def save_sorted_file(self):
        lang = self.lang.get()
        if self.df_sorted is None:
            messagebox.showwarning(LANGS[lang]["no_data"], LANGS[lang]["please_load"])
            return
        output_path = filedialog.asksaveasfilename(
            title=LANGS[lang]["save_file"],
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not output_path:
            return
        try:
            self.df_sorted.to_excel(output_path, index=False, header=self.header_var.get())

            wrap_col_name = self.wrap_col_var.get() if self.wrap_col_var.get() else None
            autosize_and_wrap_column(output_path, wrap_col_name=wrap_col_name, wrap_width=40)
            messagebox.showinfo(LANGS[lang]["success"], LANGS[lang]["saved"].format(output_path))

        except Exception as e:
            messagebox.showerror(LANGS[lang]["error"], LANGS[lang]["could_not_save"].format(e))

    def cancel_selection(self):
        self.file_path = None
        self.df = None
        self.df_sorted = None
        self.sheet_names = []
        self.sheet_var.set("")
        self.sheet_combo['values'] = []
        self.sort_col_var.set("0")
        self.sort_col_combo['values'] = []
        self.wrap_col_var.set("")
        self.wrap_col_combo['values'] = []
        self.header_var.set(0)
        self.widgets['save'].config(state="disabled")
        self.file_label.config(text="")
        self.text_preview.config(state="normal")
        self.text_preview.delete(1.0, tk.END)
        self.text_preview.config(state="disabled")
        self.text_sorted.config(state="normal")
        self.text_sorted.delete(1.0, tk.END)
        self.text_sorted.config(state="disabled")

if __name__ == "__main__":
    root = tk.Tk()
    root.attributes("-alpha", 0.93)
    app = ExcelSorterApp(root)
    root.mainloop()